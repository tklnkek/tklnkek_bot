from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, CallbackQueryHandler, MessageHandler, filters, ContextTypes
import logging
import os
from openpyxl import Workbook, load_workbook
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from flask import Flask
from threading import Thread
# Настройка логирования
logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO)
logger = logging.getLogger(__name__)

# Настройка Flask для Web Service
app = Flask(__name__)

@app.route('/')
def home():
    return "Bot is running!"

def run_flask():
    port = int(os.getenv('PORT', 8080))
    app.run(host='0.0.0.0', port=port)

logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO)

logger = logging.getLogger(__name__)
#настройка Google Sheets API
SCOPE = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
CREDENTIALS_FILE = 'tklnkekbot-839a7db9f97d.json'
SPREADSHEET_ID = '1B32QZXrYurwh964eUHxSW3J_AAZ6j2-8eOhAv-aX4G8'
SHEET_NAME = 'Лист1'
SERVICES = """
**Наши услуги:**
1. **Разработка Telegram-ботов под ключ**:
   - Автоматизация заявок, рассылок, FAQ, квизов
   - Воронки, формы, CRM-интеграции
2. **Создание Mini Apps (встроенных приложений в Telegram)**:
   - Интерфейс с кнопками, формами, каталогами
   - Подключение к API, базам данных, платёжным системам
3. **Сопровождение и доработка ботов**:
   - Поддержка существующих решений
   - Рефакторинг, добавление новых функций
   - Оптимизация скорости
4. **Консультации и проектирование**:
   - Поможем спроектировать логику бота от А до Я под вашу задачу
   - Оценим сложность, сроки, подскажем лучшие практики
"""
MAIN_KEYBOARD = InlineKeyboardMarkup([
    [InlineKeyboardButton("Услуги", callback_data='services')],
    [InlineKeyboardButton("Оставить заявку", callback_data='apply')]
])
SERVICE_KEYBOARD = InlineKeyboardMarkup([
    [InlineKeyboardButton("1. Разработка ботов", callback_data='1')],
    [InlineKeyboardButton("2. Mini Apps", callback_data='2')],
    [InlineKeyboardButton("3. Сопровождение", callback_data='3')],
    [InlineKeyboardButton("4. Консультации", callback_data='4')]
])

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):

    await update.message.reply_text(
        "Привет! Это твой личный ассистент.\n"
        "Я помогу тебе выбрать услугу и передам заявку нашей команде.\n"
        "Нажми «Услуги», чтобы посмотреть, что мы предлагаем, или выбери действие из меню ниже.",
        reply_markup=MAIN_KEYBOARD
    )


async def button(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == 'services':
        await query.message.reply_text(SERVICES, parse_mode='Markdown', reply_markup=MAIN_KEYBOARD)
    elif query.data == 'apply':
        context.user_data['apply_step'] = 'name'
        await query.message.reply_text("Пожалуйста, введите ваше имя:")
    elif query.data in ['1', '2', '3', '4']:
        user_data = context.user_data
        if user_data.get('apply_step') == 'service':
            user_data['service'] = query.data
            # словарь для преобразования номера в название услуги
            service_map = {
                '1': 'Разработка Telegram-ботов под ключ',
                '2': 'Создание Mini Apps',
                '3': 'Сопровождение и доработка ботов',
                '4': 'Консультации и проектирование'
            }
            # текст заявки
            application = (
                f"Новая заявка:\n"
                f"Имя: {user_data['name']}\n"
                f"Телефон: {user_data['phone']}\n"
                f"Услуга: {service_map[query.data]}"
            )

            await query.message.reply_text("Спасибо за заявку! Наша команда скоро свяжется с вами.", reply_markup=MAIN_KEYBOARD)

            logger.info(application)#лог заявки в терминал но можно добавить запись в гугл таблицу какую-нибудь или в личку ->
            #запись заявки в текстовый файлик
            with open('applications.txt', 'a', encoding='utf-8') as f:
                f.write(application + '\n---\n')
            #запись заявочки в табличку
            excel_file = 'applications.xlsx'
            if not os.path.exists(excel_file):
                wb = Workbook()
                ws = wb.active
                ws.title = 'Applications'
                ws.append(['Имя', 'Телефон', 'Услуга', 'Время'])
                wb.save(excel_file)
            wb = load_workbook(excel_file)
            ws = wb.active
            ws.append([
                user_data['name'],
                user_data['phone'],
                service_map[query.data],
                query.message.date.strftime('%Y-%m-%d %H:%M:%S')
            ])
            wb.save(excel_file)
            #запись в гугл табличку
            try:
                credentials = ServiceAccountCredentials.from_json_keyfile_name(CREDENTIALS_FILE, SCOPE)
                client = gspread.authorize(credentials)
                sheet = client.open_by_key(SPREADSHEET_ID).worksheet(SHEET_NAME)
                sheet.append_row([
                    user_data['name'],
                    user_data['phone'],
                    service_map[query.data],
                    query.message.date.strftime('%Y-%m-%d %H:%M:%S')
                ])
            except Exception as e:
                logger.error(f"ошибка записи в гугл табличку: {e}")
            user_data.clear()

#обработчик текстовых сообщений для заявки
async def handle_application(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_data = context.user_data
    #текст отправленный юзером
    text = update.message.text
    if user_data.get('apply_step') == 'name':
        user_data['name'] = text
        user_data['apply_step'] = 'phone'
        await update.message.reply_text("Введите ваш номер телефона:")
    elif user_data.get('apply_step') == 'phone':
        user_data['phone'] = text
        user_data['apply_step'] = 'service'
        keyboard = [
            [InlineKeyboardButton("1. Разработка ботов", callback_data='1')],
            [InlineKeyboardButton("2. Mini Apps", callback_data='2')],
            [InlineKeyboardButton("3. Сопровождение", callback_data='3')],
            [InlineKeyboardButton("4. Консультации", callback_data='4')]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await update.message.reply_text(
            "Выберите услугу, нажав на кнопку:\n" + SERVICES,
            reply_markup=reply_markup,
            parse_mode='Markdown'
        )
    elif user_data.get('apply_step') == 'service':
        await update.message.reply_text(
            "Пожалуйста, выберите услугу, нажав на кнопку:\n" + SERVICES,
            reply_markup=SERVICE_KEYBOARD,
            parse_mode='Markdown'
        )

async def error_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    logger.error(f"Update {update} caused error {context.error}")

def main():
    flask_thread = Thread(target=run_flask)
    flask_thread.start()

    TOKEN = os.getenv('TELEGRAM_TOKEN')
    application = Application.builder().token(TOKEN).build()
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CallbackQueryHandler(button))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_application))
    application.add_error_handler(error_handler)
    application.run_polling()


if __name__ == '__main__':
    main()
